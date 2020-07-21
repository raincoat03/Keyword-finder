from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
start = time.time()
date = datetime.today()

chromedriver = "C:/Users/june/Desktop/chromedriver.exe"
options = webdriver.ChromeOptions()
'''
options.add_argument('headless')
options.add_argument('disable-gpu')
options.add_argument('window-size=1920x1080')
options.add_argument('lang=ko_KR')
options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36")
'''
driver = webdriver.Chrome(chromedriver, options = options)
url_naver = "https://searchad.naver.com/login"
id_naver = "lyrical98"
password_naver = "dmsgur!23"
search_total = []
naver_total = []
already_keyword_list = []

wb = openpyxl.Workbook()
wb.save("keyword.xlsx")
sheet = wb.active
sheet["A2"] = "keyword"
sheet["B2"] = "M+PC 클릭수"
sheet["C2"] = "예상 클릭수"
sheet["D2"] = "입찰가"
sheet["E2"] = "예상비용"

workbook = load_workbook("C:/Users/june/Desktop/keyword/keyword.xlsx", data_only=True)
worksheet = workbook["Sheet"]
cell_range = worksheet["A3":"A90000"]
for i in cell_range:
    for j in i:
        already_keyword_list.append(j.value)
keyword = [v for v in already_keyword_list if v]



driver.get(url_naver)
id_xpath = "/html/body/marvel-root/login/div/div/div/div/fieldset/dl/dd[1]/input"
password_xpath = "/html/body/marvel-root/login/div/div/div/div/fieldset/dl/dd[2]/input"
naver_id_login = driver.find_element_by_xpath(id_xpath)
naver_id_login.send_keys(id_naver)
naver_password_login = driver.find_element_by_xpath(password_xpath)
naver_password_login.send_keys(password_naver)
naver_login_button_xpath = "/html/body/marvel-root/login/div/div/div/div/fieldset/div/span/button"
driver.find_element_by_xpath(naver_login_button_xpath).click()

time.sleep(3)

main = driver.window_handles
for handle in main:
    if handle != main[0]:
        driver.switch_to.window(handle)
        driver.close()
driver.switch_to.window(driver.window_handles[0])
naver_class_n = ""
keyword_naver_url = "https://manage.searchad.naver.com/customers/1948785/tool/keyword-planner?keywords="

for i in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 18, 19, 20, 21, 22]:
    keyword_naver_url = "https://manage.searchad.naver.com/customers/1948785/tool/keyword-planner?keywords="
    time.sleep(2)
    driver.get(keyword_naver_url)
    element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-planner/div[2]/div[1]/div[1]/div[2]/div/div/div[2]/div[2]/label/div/ul/li/i")))
    driver.execute_script("arguments[0].click();", element)
    time.sleep(1)
    element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH,"/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-planner/div[2]/div[1]/div[1]/div[2]/div/div/div[2]/div[2]/label")))
    driver.execute_script("arguments[0].click();", element)
    time.sleep(2)
    naver_class_n = str(i)
    naver_class_1 = "/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-planner/div[2]/div[1]/div[1]/div[2]/div/div/div[2]/div[2]/div/div/li["
    naver_class_2 = naver_class_n
    naver_class_3 = "]/div"  # 3 ~ 22
    click_find = naver_class_1 + naver_class_2 + naver_class_3
    time.sleep(1)
    element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH,"/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-planner/div[2]/div[1]/div[1]/div[2]/div/div/div[2]/div[2]/div/button")))
    driver.execute_script("arguments[0].click();", element)
    driver.find_element_by_xpath(click_find).click()
    time.sleep(1)
    element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH,"/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-planner/div[2]/div[1]/div[1]/div[3]/button")))
    driver.execute_script("arguments[0].click();", element)
    time.sleep(1)
    ## 검색어 추출
    for j in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
        search_name_list, search_pc_list, search_mobile_list = [], [], []
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        search_name = soup.select("tr > td > elena-keyword > span")
        for i in search_name:
            i = i.text.strip()
            if i[-1] == "S":
                i.rstrip("S")
            search_name_list.append(i)
            if "일부노출" in search_name_list:
                search_name_list.remove("일부노출")
            if "S" in search_name_list:
                search_name_list.remove("S")
            if "19" in search_name_list:
                search_name_list.remove("19")

        search_click_pc = soup.select("td.elenaColumn-monthlyAvePcClkCnt")
        for k in search_click_pc:
            k = k.text.strip()
            k = k.replace(",","")
            search_pc_list.append(k)

        search_click_mobile = soup.select("td.elenaColumn-monthlyAveMobileClkCnt")
        for l in search_click_mobile:
            l = l.text.strip()
            l = l.replace(",","")
            search_mobile_list.append(l)

        for i in range(len(search_name_list)):
            sum = int(round(float(search_pc_list[i]),-1))+int(round(float(search_mobile_list[i]),-1))
            temp = [search_name_list[i], sum]
            search_total.append(temp)

        next_page_1 = "/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-planner/div[2]/div[1]/div[2]/div[3]/elena-table/elena-table-paginator/div/div/nav/ul/li["
        next_page_2 = str(j)
        next_page_3 = "]/a"
        next_page = next_page_1 + next_page_2 + next_page_3
        time.sleep(1)
        driver.find_element_by_xpath(next_page).click()
        time.sleep(1)

search_total.sort(key=lambda x:-x[1])  # 클릭수 높은 순서대로 정렬
print("수집완료")

# 목록 입력
for i in range(len(search_total)):
    if search_total[i][0] not in already_keyword_list:
        driver.get(keyword_naver_url)
        time.sleep(1)
        keyword_enter_sheet = "/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-planner/div[2]/div[2]/div/div/form/div[1]/textarea"
        element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-planner/div[2]/div[2]/div/div/form/div[1]/textarea")))
        driver.execute_script("arguments[0].click();", element)
        keyword_enter = driver.find_element_by_xpath(keyword_enter_sheet)
        time.sleep(1)
        keyword_enter.send_keys(search_total[i][0])
        time.sleep(1)
        driver.find_element_by_xpath("/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-planner/div[2]/div[2]/div/div/form/div[2]/button[2]").click()
        time.sleep(2)
        bid_price = driver.find_element_by_xpath("/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-estimate/div[3]/div[2]/div/div[1]/div[1]/div/elena-input-amt/div/input")
        for j in range(70, 100001, 50):
            time.sleep(1)
            element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-estimate/div[3]/div[2]/div/div[1]/div[1]/div/elena-input-amt/div/input")))
            driver.execute_script("arguments[0].click();", element) # bid_price_empty
            n = j
            bid_price.clear()
            bid_price.send_keys(j)
            driver.find_element_by_xpath("/html/body/elena-root/elena-wrap/div/div[2]/elena-tool-wrap/div/div/div/div/elena-keyword-estimate/div[3]/div[2]/div/div[1]/div[4]/button").click()
            time.sleep(2)
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            click = soup.select("td.elenaColumn-clicks")
            cost = soup.select("td.elenaColumn-cost")
            cost_list = []
            for j in click:
                j = j.text.strip()
                j = j.replace(",", "")
                j = int(j)
                expect_click = j
            for k in cost:
                k = k.text.strip()
                k = k.strip("원")
                k = k.replace(",", "")
                k = k.strip()
                k = int(k)
                cost_list.append(k)
            if j >= search_total[i][1] + 30:
                naver_total.append([search_total[i][0],search_total[i][1],j,n,cost_list[0]])
                break
            time.sleep(1)
        sheet.append(naver_total[-1])
print(time.time()-start)